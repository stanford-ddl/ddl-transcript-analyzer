import os
import pandas as pd
import util
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from config import IS_DEBUG, TOTAL_SESSIONS, IS_ANALYZE_ALL_SESSIONS, session_num, DATA_DIR, RESULTS_DIR, PROCESSING_DIR

# cleaning arguments
def concat_args(arguments, deliberation):
  cleaned_args = {deliberation: []}
  for arg in arguments:
    args = arg[0].split("\n")
    joined = " ".join([arg[3:] for arg in args])
    cleaned_args[deliberation].append((joined, arg[1]))
  return cleaned_args

# collecting arguments in a deliberation
def extract_args(path, deliberation):
    df = pd.DataFrame(pd.read_excel(path))
    start_col = df.columns.get_loc("All Arguments Summarized")
    cols = df.columns[start_col:]

    # filtering out debugging columns
    filter_cols = [col for col in cols if col.find("debugging") == -1]
    relevant_args_df = df.dropna(subset=filter_cols, how='all')

    # list of Fileread's summarized arguments and their indexed order
    arguments = list((relevant_args_df["All Arguments Summarized"], relevant_args_df["Order"]))
    arguments = [(x, y) for x, y in zip(arguments[0], arguments[1]) if str(x) != 'nan']
    return concat_args(arguments, deliberation)

# Given a Worksheet,
# wrap all text cells.
def wrap_text(ws):
   for row in ws.iter_rows():
      for cell in row:
         cell.alignment = Alignment(wrap_text=True)

# Given a Worksheet,
# resize all columns so they can reasonably hold their information
def resize_columns(ws):
   BUFFER = 10
   for col in ws.columns:
      width = 0
      column = col[0].column_letter
      for cell in col:
         try:
            if len(str(cell.value)) > width:
               width = len(str(cell.value))
         except:
            pass
      ws.column_dimensions[column].width = width / 2 + BUFFER

# this function takes in text from a deliberation and determines if it contains arguments
def check_arguments(text):
    system_prompt = "Determine if the following text contains arguments. Respond ONLY with 'yes' or 'no'. Do not provide an explanation, do not capitalize Yes or No, do not put any punctuation, only respond 'yes' or 'no'."
    response = util.simple_llm_call(system_prompt, text)
    # if invalid response, try again
    if response not in ['yes', 'no']:
        # print("bad output")
        response = util.simple_llm_call(system_prompt, text)
    # print(text + "----->" + response)
    return response.strip().lower() == 'yes'

# Given a sheet,
# format it appropriately.
def format_sheet(ws):
  # If this sheet is from the "Step 1: Downloaded CSVs" Google Drive
  if ws.cell(row=1, column=1).value == "roomId":
     ws.delete_rows(1)
     ws.delete_cols(1)
     ws.delete_cols(2, 2)
  
  # If this sheet is from the "Step 3: Fileread Results" Google Drive
  if ws.cell(row=1, column=1).value == "Order":
     ws.delete_cols(1)
  NUM_COLUMNS_TO_KEEP = 3
  if ws.max_column > NUM_COLUMNS_TO_KEEP:
     for col in range(ws.max_column, NUM_COLUMNS_TO_KEEP, -1):
        ws.delete_cols(col)
  ws.cell(row=1, column=1).value = "Speaker"
  ws.cell(row=1, column=2).value = "Time"
  ws.cell(row=1, column=3).value = "Text"

  for row in ws.iter_rows():
     for cell in row:
        cell.style = "Normal"

# this function duplicates the input spreadsheet in destination folder, and creates and populates contians args column
def create_args_sheet(file_path, destination_folder):
    wb = load_workbook(file_path)
    ws = wb.worksheets[0]

    file_name = os.path.basename(file_path)
    new_file_path = os.path.join(destination_folder, file_name)

    if os.path.exists(new_file_path):
       print("Skipped", file_name, "because it has previously been processed")
       return

    format_sheet(ws) # Format the file
    
    # Define columns
    ORDER_COL = 1 # Order
    TEXT_COL = 4 # text
    HAS_ARG_COL = 5 # Has Arguments
    ARG_SUM_COL = 6 # All Arguments Summarized

    # Add the new columns
    ws.insert_cols(1) # Insert "Order" column
    ws.cell(row=1, column=ORDER_COL, value="Order")
    ws.cell(row=1, column=HAS_ARG_COL, value="Has Arguments")
    ws.cell(row=1, column=ARG_SUM_COL, value="All Arguments Summarized")

    # Iterate over the rows and populate the "has arguments" column
    for row in range(2, ws.max_row + 1):
        text = ws.cell(row=row, column=TEXT_COL).value
        if text is not None:  # Check if the cell is not empty
            has_arguments = check_arguments(text)
            ws.cell(row=row, column=HAS_ARG_COL, value="yes" if has_arguments else "no")
        else:
            ws.cell(row=row, column=HAS_ARG_COL, value="no")
        # Construct the "Order" column
        ws.cell(row=row, column=ORDER_COL, value=row-1)
    
    resize_columns(ws)

    # Iterate over the rows again to populate the "all arguments summarized" column
    for row in range(2, ws.max_row + 1):
        has_arguments = ws.cell(row=row, column=HAS_ARG_COL).value
        if has_arguments == "yes":
            text = ws.cell(row=row, column=TEXT_COL).value
            # summarized_text = util.simple_llm_call("Summarize the following text in a numbered list. For example, the text input ' thank you. so the last thing was that if there's ten candidates and you only pick your top two, then your and they don't get it, your vote doesn't count it all. so the best thing to do is you got to rank all 10, but again, not enough data on the subject to make predictions. and i think mike use' should be summarized as '1. If you only pick your top two candidates out of ten, your vote doesn't count if they don't win. 2. The best approach is to rank all ten candidates. 3. There is not enough data available to make predictions on the subject.' Follow this formatting exactly.", text)
            summarized_text = util.simple_llm_call("Summarize the following text in a numbered list.  Follow the output format '1. argument 1 \n 2. argument 2 \n' etcetera. Every argument MUST be on a different line. There could be one or more arguments.", text)
            ws.cell(row=row, column=ARG_SUM_COL).value = summarized_text
            #print("+++++++++" + text)
            #print("==========" + summarized_text)

    wrap_text(ws)
    
    # Save the modified file
    wb.save(new_file_path)
    print("Processed", file_name)

# Given a path for a 'processing' folder,
# create that folder if needed.
def create_processing_path(processing_path):
   print("\nCreating Session", session_num, "processing folder...", end=" ")
   # Creates the required results folder if it does not already exist
   os.makedirs(processing_path, exist_ok=True)
   print("Done")

# Code starts here
# Given a data folder and empty argument variables,
# process all of the data and find arguments in the text.
def process_cleaned_data(data_path, all_args_indexed, all_args):
    processing_path = os.path.join(PROCESSING_DIR, session_num)
    create_processing_path(processing_path)
    print("\nIdentifying arguments in Session", session_num + "...")
    # looping over all deliberations and collecting 1) the arguments presented and 2) the index of each argument in that deliberation
    for deliberation in os.listdir(data_path):
      path = os.path.join(data_path, deliberation)
      if path.endswith('xlsx') or path.endswith('csv'):
        create_args_sheet(path, processing_path)
        new_path = os.path.join(processing_path, deliberation)
        formatted_args = extract_args(new_path, deliberation)
        all_args_indexed.update(formatted_args)
        all_args += all_args_indexed[deliberation]
    print("Finished identifying arguments in Session", session_num)